# excel_splitter.py
"""
Excel Splitting Backend
-----------------------
• Pure logic only – no GUI
• Reads an input .xlsx, reshapes the columns and
  – fills Unit_Price = Total_Line_Value / Quantity (2-dec round)
  – bumps any Total_Line_Value < 1.00 up to 1.00
• Writes out ≤499-row chunks with a header template
• Saves <MAWB>_adjusted_rows.xlsx for rows that were bumped
  (values shown there are the *original* numbers, before bumping)
"""

import os, re, string
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

import datetime

# ─────────────── constants ────────────────────────────────────
APP_DIR        = Path(__file__).resolve().parent
HEADER_PATH    = APP_DIR / "Resources" / "ExcelSplitter" / "Header Sample.xlsx"
ROWS_PER_FILE  = 495

HEADERS = [
    "Invoice_No","Part","Commercial_Description","Country_of_Origin","Country_of_Export",
    "Tariff_Number","Quantity","Quantity_UOM","Unit_Price","Total_Line_Value",
    "Net_Weight_KG","Gross_Weight_KG","Manufacturer_Name","Manufacturer_Address_1",
    "Manufacturer_Address_2","Manufacturer_City","Manufacturer_State",
    "Manufacturer_Zip","Manufacturer_Country","MID_Code","Buyer_Name",
    "Buyer_Address_1","Buyer_Address_2","Buyer_City","Buyer_State","Buyer_Zip",
    "Buyer_Country","Buyer_ID_Number","Consignee_Name","Consignee_Address_1",
    "Consignee_Address_2","Consignee_City","Consignee_State","Consignee_Zip",
    "Consignee_Country","Consignee_ID_Number","SICountry","SP1","SP2",
    "Zone_Status","Privileged_Filing_Date","Line_Piece_Count","ADD_Case_Number",
    "CVD_Case_Number","AD_Non_Reimbursement_Statement",
    "AD-CVD_Certification_Designation",
]

# raw-column → target-column (letter to letter)
MAPPING = {
    "B": "F", "D": "G", "F": "J", "E": "L",
    "G": "M", "H": "N", "I": "P", "K": "R", "M": "T",
}

# hard-coded fills
CONSTANTS = {
    "D": "CN", "E": "CN", "S": "CN", "H": "PCS",
}

# ───────────────── helper funcs ───────────────────────────────
def xl_idx(col: str) -> int:
    """Excel-style column letter → zero-based index."""
    idx = 0
    for ch in col.upper():
        idx = idx * 26 + (ord(ch) - 64)
    return idx - 1


def get_mawb(path: str) -> str:
    wb   = load_workbook(path, read_only=True, data_only=True)
    mawb = str(wb.active["U9"].value).strip()
    wb.close()
    return mawb


# ───────────────── dataframe prep ─────────────────────────────
def prepare_dataframe(path: str) -> pd.DataFrame:
    # read starting at row 10 (skip first 9 rows)
    raw = pd.read_excel(path, skiprows=9, engine="openpyxl")

    # detect commodity-description column (new layout)
    headers  = [str(h).lower() for h in raw.columns]
    has_desc = any("commodity" in h for h in headers)

    # build mapped dict
    mapped = {}
    if has_desc:
        # Column C (idx 2) → Commercial_Description
        mapped[HEADERS[xl_idx("C")]] = raw.iloc[:, 2]

        # shift all source cols ≥C right by one
        for src, tgt in MAPPING.items():
            src_idx  = xl_idx(src)
            real_idx = src_idx + 1 if src_idx >= 2 else src_idx
            mapped[HEADERS[xl_idx(tgt)]] = raw.iloc[:, real_idx]
    else:
        # legacy layout
        for src, tgt in MAPPING.items():
            mapped[HEADERS[xl_idx(tgt)]] = raw.iloc[:, xl_idx(src)]

    # drop rows that are completely blank
    base = pd.DataFrame(mapped).dropna(how="all").reset_index(drop=True)

    # start full frame with all headers
    df = pd.DataFrame(index=base.index, columns=HEADERS)
    df.update(base)

    # inject constants
    for tgt, val in CONSTANTS.items():
        df[HEADERS[xl_idx(tgt)]] = val

    # SG / HK country overrides based on MID prefix
    mid_c, cntry_c = HEADERS[xl_idx("T")], HEADERS[xl_idx("S")]
    sg_mask = df[mid_c].astype(str).str.upper().str.startswith("SG", na=False)
    hk_mask = df[mid_c].astype(str).str.upper().str.startswith("HK", na=False)
    df.loc[sg_mask, cntry_c] = "SG"
    df.loc[hk_mask, cntry_c] = "HK"

    # pad ZIP codes to 6 digits
    def _pad_zip(x):
        if pd.isna(x):
            return ""
        s = str(x).strip()
        m = re.fullmatch(r'(\d+)(?:\.0*)?', s)
        return f"{int(m.group(1)):06d}" if m else s

    for zip_col in ("Manufacturer_Zip", "Buyer_Zip"):
        if zip_col in df.columns:
            df[zip_col] = df[zip_col].apply(_pad_zip)

    # ─────── calculate Unit_Price  (J / G) ───────
    qty_col   = HEADERS[xl_idx("G")]   # Quantity
    total_col = HEADERS[xl_idx("J")]   # Total_Line_Value
    unit_col  = HEADERS[xl_idx("I")]   # Unit_Price

    df[unit_col] = (
        pd.to_numeric(df[total_col], errors="coerce") /
        pd.to_numeric(df[qty_col],   errors="coerce")
    ).round(2)

    return df


# ───────────────── save chunks to disk ────────────────────────
def save_chunks(
    df: pd.DataFrame,
    out_dir: str | Path,
    mawb: str,
    rows: int = ROWS_PER_FILE
) -> int:
    out_dir = Path(out_dir)

    # ── build & create the MAWB-specific folder ──────────────────
    date_str = datetime.date.today().strftime("%Y-%m-%d")
    sub_dir  = out_dir / f"GA_CI_{mawb}_{date_str}"
    sub_dir.mkdir(parents=True, exist_ok=True)

    # choose suffix list based on rows-per-file
    if rows < 600:
        part_list = [f"{ltr}{i+1}" for ltr in string.ascii_uppercase for i in range(2)]
    else:
        part_list = list(string.ascii_uppercase)

    if len(df) > rows * len(part_list):
        raise ValueError("Too many rows for available file parts.")

    # header template from sample file
    template_headers = [
        cell.value
        for cell in load_workbook(HEADER_PATH, read_only=True).active[1]
        if cell.value
    ]

    # quick references
    total_col = HEADERS[xl_idx("J")]   # Total_Line_Value
    qty_col   = HEADERS[xl_idx("G")]   # Quantity
    unit_col  = HEADERS[xl_idx("I")]   # Unit_Price

    adj_rows: list[pd.DataFrame] = []   # collect original rows we bump
    part = 0

    for start in range(0, len(df), rows):
        chunk   = df.iloc[start : start + rows].copy()
        suffix  = part_list[part]
        invoice = f"{mawb}-{suffix}"
        chunk["Invoice_No"] = invoice

        # ─────── bump any Total_Line_Value < 0.51 ───────
        mask = pd.to_numeric(chunk[total_col], errors="coerce") < 0.51
        if mask.any():
            adj_rows.append(chunk.loc[mask].copy())      # keep originals
            chunk.loc[mask, total_col] = 0.51            # bump
            chunk.loc[mask, unit_col] = (
                pd.to_numeric(chunk.loc[mask, total_col], errors="coerce") /
                pd.to_numeric(chunk.loc[mask, qty_col],   errors="coerce")
            ).round(2)

        # ─────── write this split workbook ───────
        chunk = chunk.reindex(columns=template_headers)
        file_name = f"GA_CI_{invoice}_{date_str}.xlsx"
        xlsx_path = sub_dir / file_name

        with pd.ExcelWriter(xlsx_path, engine="xlsxwriter") as writer:
            chunk.to_excel(writer, index=False, header=False, startrow=1)
            wb  = writer.book
            ws  = writer.sheets["Sheet1"]
            fmt = wb.add_format({
                "bold": True, "font_name": "Times New Roman",
                "font_size": 12, "align": "center", "valign": "center",
            })
            for idx, title in enumerate(template_headers):
                ws.write(0, idx, title, fmt)
            ws.set_column(0, 0, max(len(invoice), len("Invoice_No")) + 2)
            ws.set_column(5, 5, 20)

        part += 1

    # ─────── write adjustment-log workbook ───────
    if adj_rows:
        adj_df = pd.concat(adj_rows, ignore_index=True)
        adj_df = adj_df.reindex(columns=template_headers)
        adj_df[unit_col] = (
            pd.to_numeric(adj_df[total_col], errors="coerce") /
            pd.to_numeric(adj_df[qty_col],   errors="coerce")
        ).round(2)

        adj_path = sub_dir / f"{mawb}_adjusted_rows.xlsx"
        adj_df.to_excel(adj_path, index=False)

    return part