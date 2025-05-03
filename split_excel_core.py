"""
Backend engine
Splits the workbook into 998-row chunks, applies column mapping / constants,
drops fully-blank rows, widens column F, and saves files named <MAWB>-A.xlsx …
"""


import os, sys
from math import ceil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, PatternFill, Alignment

ROWS_PER_FILE = 998

MAPPING = {
    "B": "F",
    "D": "G",
    "F": "J",
    "E": "L",
    "G": "M",
    "H": "N",
    "I": "P",
    "K": "R",
    "M": "T"
}

CONSTANTS = {
    "D": "CN",          # Country_of_Origin
    "E": "CN",          # Country_of_Export
    "S": "CN",          # Manufacturer_Country
    "H": "PCS"
}

HEADERS = [
    "Invoice_No","Part","Commercial_Description","Country_of_Origin","Country_of_Export",
    "Tariff_Number","Quantity","Quantity_UOM","Unit_Price","Total_Line_Value",
    "Net_Weight_KG","Gross_Weight_KG","Manufacturer_Name","Manufacturer_Address_1",
    "Manufacturer_Address_2","Manufacturer_City","Manufacturer_State","Manufacturer_Zip",
    "Manufacturer_Country","MID_Code","Buyer_Name","Buyer_Address_1","Buyer_Address_2",
    "Buyer_City","Buyer_State","Buyer_Zip","Buyer_Country","Buyer_ID_Number",
    "Consignee_Name","Consignee_Address_1","Consignee_Address_2","Consignee_City",
    "Consignee_State","Consignee_Zip","Consignee_Country","Consignee_ID_Number",
    "SICountry","SP1","SP2","Zone_Status","Privileged_Filing_Date","Line_Piece_Count",
    "ADD_Case_Number","CVD_Case_Number","AD_Non_Reimbursement_Statement",
    "AD-CVD_Certification_Designation",
]

# ───────────────── helpers ─────────────────
def xl_idx(col: str) -> int:
    idx = 0
    for ch in col.upper():
        idx = idx*26 + (ord(ch) - 64)
    return idx - 1

def get_mawb(src_path: str) -> str:
    wb = load_workbook(src_path, read_only=True, data_only=True)
    mawb = str(wb.active["U9"].value).strip()
    wb.close()
    return mawb

def build_dataframe(src_path: str) -> pd.DataFrame:
    """
    1. Read the sheet (skip first 9 rows).
    2. Pull mapped columns.
    3. Drop rows that are completely empty **before** adding constants.
    4. Re-insert constants and expand to the full 46-column layout.
    """
    raw = pd.read_excel(src_path, skiprows=9, engine="openpyxl")

    # 1 ─ pull mapped columns into a dict keyed by *header names*
    mapped = {}
    for src_letter, tgt_letter in MAPPING.items():
        header_name = HEADERS[xl_idx(tgt_letter)]   # e.g. "Tariff_Number"
        mapped[header_name] = raw.iloc[:, xl_idx(src_letter)]

    # 2 ─ build interim DataFrame, drop fully-blank rows
    tmp_df = pd.DataFrame(mapped).dropna(how="all").reset_index(drop=True)

    # 3 ─ start the final 46-column frame (all headers present, empty)
    final_df = pd.DataFrame(index=tmp_df.index, columns=HEADERS)

    # 4 ─ copy mapped data into their proper header columns
    for col in tmp_df.columns:
        final_df[col] = tmp_df[col]

    # 5 ─ add constant columns (again, by header name)
    for tgt_letter, value in CONSTANTS.items():
        header_name = HEADERS[xl_idx(tgt_letter)]
        final_df[header_name] = value

    return final_df



def save_chunks(df: pd.DataFrame, out_dir: str, mawb: str, rows: int = 998) -> int:
    import string

    os.makedirs(out_dir, exist_ok=True)

    part_list = [f"{letter}{i+1}" for letter in string.ascii_uppercase for i in range(3)]

    # Load header sample to extract correct columns
    sample_path = "./RealData/Header Sample.xlsx"
    sample_wb = load_workbook(sample_path, read_only=True)
    sample_ws = sample_wb.active
    template_headers = [cell.value for cell in sample_ws[1] if cell.value]
    sample_wb.close()

    master_txt = os.path.join(out_dir, f"MID_report_{mawb}.txt")
    with open(master_txt, "w", encoding="utf-8") as rpt:
        rpt.write("File\tRow\tOriginalMID\tFinalMID\tManufacturer\tAddress\n")

        part = 0
        for start in range(0, len(df), rows):
            chunk = df.iloc[start:start + rows].copy()
            suffix = part_list[part]
            invoice = f"{mawb}-{suffix}"
            chunk["Invoice_No"] = invoice

            # Fix MIDs
            chunk = mid_fixer.fix_mids(chunk, mawb, suffix, out_dir)

            # Drop helper columns
            for col in ["MID_flag", "MID_original"]:
                if col in chunk.columns:
                    chunk.drop(columns=col, inplace=True)

            # Match template header exactly
            chunk = chunk.reindex(columns=template_headers)

            # Report modified MIDs
            for r, (_, row) in enumerate(chunk.iterrows(), start=2):
                if "MID_flag" in row and row["MID_flag"]:
                    rpt.write(f"{invoice}.xlsx\t{r}\t{row.get('MID_original', '')}\t"
                              f"{row['MID_Code']}\t{row['Manufacturer_Name']}\t"
                              f"{row['Manufacturer_Address_1']}\n")

            # Write to Excel using xlsxwriter
            xlsx_path = os.path.join(out_dir, f"{invoice}.xlsx")
            with pd.ExcelWriter(xlsx_path, engine='xlsxwriter') as writer:
                # Write without header at row 1
                chunk.to_excel(writer, index=False, header=False, startrow=1)

                # Write header manually at row 0
                workbook  = writer.book
                worksheet = writer.sheets['Sheet1']
                header_format = workbook.add_format({
                    'bold': True,
                    'font_name': 'Times New Roman',
                    'font_size': 12,
                    'align': 'center',
                    'valign': 'vcenter'
                })

                for col_idx, col_name in enumerate(template_headers):
                    worksheet.write(0, col_idx, col_name, header_format)

                # Optional: Adjust column widths for specific columns
                worksheet.set_column(0, 0, max(len(invoice), len("Invoice_No")) + 2)  # Invoice_No
                worksheet.set_column(5, 5, 20)  # Column F

            part += 1

    return part

# optional CLI
if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python split_excel_core.py <input.xlsx> <output_folder>")
        sys.exit(1)

    src, dst = sys.argv[1], sys.argv[2]
    parts = save_chunks(build_dataframe(src), dst, get_mawb(src))
    print(f"Done – {parts} file(s) written to '{dst}'.")
